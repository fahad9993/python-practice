import openpyxl
import tkinter as tk


def show_definition(event):
    global definition_label  # Make definition_label accessible within the function
    # Get the cell coordinates
    col = event.widget.grid_info()['column']
    row = event.widget.grid_info()['row']

    # Check if the hovered column is "Word" column
    if col == 1:  # Assuming "Word" column is the second column (index 1)
        # Get the word from the hovered cell
        word = sheet[event.widget.grid_info()['row']][1].value  # Assuming "Word" column is the second column (index 1)

        # Initialize definition with an empty string
        definition = ""

        # Find the corresponding definition in the same row
        for row_data in sheet.iter_rows(values_only=True):
            if row_data[0] == word:
                definition = row_data[3]  # Assuming "Definition" column is the fourth column (index 3)
                break

        # Update the definition label text
        definition_label.config(text=definition)



# Provide the file path of the Excel file
excel_file_path = 'D:\Study\GRE\A3.xlsx'

# Load Excel file
workbook = openpyxl.load_workbook(excel_file_path)

# Create GUI
root = tk.Tk()
root.title("Excel Hover-over Pop-up")

# Create pop-up window for definition
popup = tk.Toplevel(root)
popup.title("Definition")
definition_label = tk.Label(popup, text="", font=("Helvetica", 12))
definition_label.pack()

# Iterate over sheets in the Excel file
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    label = tk.Label(root, text=sheet_name, font=("Helvetica", 16, "bold"))
    label.grid(sticky="ew")  # Use grid for the sheet name labels

    # Display Excel data in a grid
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        for j, cell in enumerate(row):
            label = tk.Label(root, text=cell, relief=tk.RIDGE, borderwidth=1, width=20, height=2)
            label.grid(row=i + len(workbook.sheetnames) + 1, column=j, sticky="nsew")  # Use grid for cell data
            label.bind("<Enter>", show_definition)  # Bind hover-over event

# Configure row and column weights for grid resizing
for i in range(len(workbook.sheetnames) + 1):
    root.grid_rowconfigure(i, weight=1)
for j in range(sheet.max_column):
    root.grid_columnconfigure(j, weight=1)

root.mainloop()
